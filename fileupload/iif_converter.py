import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import re

def convert_excel_to_iif(excel_file, output_iif_path):
    try:
        df = pd.read_excel(excel_file, engine='openpyxl', skiprows=12)
    except Exception as e:
        return f"Error reading the Excel file: {str(e)}"

    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        date_from_excel = None

        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=4, column=col).value
            if isinstance(cell_value, str):
                match = re.findall(r'(\d{2}-[A-Za-z]{3}-\d{2})', cell_value)  
                if match and len(match) >= 2:
                    date_from_excel = match[1]  
                    break  
        
        if not date_from_excel:
            raise ValueError("No valid second date found in row 4.")

        extracted_date = datetime.strptime(date_from_excel, '%d-%b-%y')  
        
        adjusted_date = extracted_date - timedelta(days=2)

        #  MM/DD/YYYY 
        formatted_date = adjusted_date.strftime('%m/%d/%Y')  
    except Exception as e:
        return f"Error extracting or adjusting the date: {str(e)}"

    df_cleaned = df[['Employee', 'Emp\nNum', 'Reg Hours', 'Total', 'Rate']]
    df_cleaned = df_cleaned.dropna(subset=['Employee', 'Emp\nNum', 'Reg Hours', 'Total'])

    iif_header = "!TIMEACT\tDATE\tJOB\tEMP\tITEM\tPITEM\tDURATION\tPROJ\tNOTE\tXFERTOPAYROLL\tBILLINGSTATUS\n"
    iif_data = [iif_header]

    for _, row in df_cleaned.iterrows():
        timeact = "TIMEACT"
        date = formatted_date    
        job = "Default Job"
        emp = row['Employee']
        emp_num = row['Emp\nNum'] 
        item = "ST Rate"  
        pitem = "Regular Pay"
        duration = row['Total'] if pd.notna(row['Total']) else row['Reg Hours']  
        proj = ""
        note = f"EMP_ID:{emp_num}"  
        xfertopayroll = "Y"
        billingstatus = "1"
        
        if duration > 80:
            regular_duration = 80  
            iif_row_regular = f"{timeact}\t{date}\t{job}\t{emp}\t{item}\t{pitem}\t{regular_duration}\t{proj}\t{note}\t{xfertopayroll}\t{billingstatus}"
            iif_data.append(iif_row_regular)
            overtime_duration = duration - 80  # Overtime 
            item = "OT Rate"
            pitem = "Overtime Pay"
            iif_row_overtime = f"{timeact}\t{date}\t{job}\t{emp}\t{item}\t{pitem}\t{overtime_duration}\t{proj}\t{note}\t{xfertopayroll}\t{billingstatus}"
            iif_data.append(iif_row_overtime)
        else:
            # If duration is 80 Reg pay
            iif_row = f"{timeact}\t{date}\t{job}\t{emp}\t{item}\t{pitem}\t{duration}\t{proj}\t{note}\t{xfertopayroll}\t{billingstatus}"
            iif_data.append(iif_row)

    iif_content = "\n".join(iif_data)

    try:
        with open(output_iif_path, 'w') as iif_file:
            iif_file.write(iif_content)
        print(f"IIF file successfully created: {output_iif_path}")
    except Exception as e:
        return f"Error saving the IIF file: {str(e)}"

    return iif_content



excel_file_path = '/mnt/data/Test_Timesheet1_MFprPOS.xlsx'  
output_iif_path = '/mnt/data/output_file.iif'  

convert_excel_to_iif(excel_file_path, output_iif_path)
